"""
Report Generator for Consolidation Pattern Analysis
Generates comprehensive reports in multiple formats
"""

import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, List, Any, Optional
import json
from pathlib import Path
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import markdown
from jinja2 import Template

logger = logging.getLogger(__name__)

class ReportGenerator:
    """Generate comprehensive analysis reports"""
    
    def __init__(self, output_dir: str = './reports'):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
    def generate_full_report(self, 
                           analysis_results: Dict,
                           format: str = 'all') -> Dict[str, str]:
        """Generate comprehensive report in specified format(s)"""
        
        generated_files = {}
        
        if format in ['json', 'all']:
            json_file = self.generate_json_report(analysis_results)
            generated_files['json'] = json_file
            
        if format in ['csv', 'all']:
            csv_files = self.generate_csv_reports(analysis_results)
            generated_files['csv'] = csv_files
            
        if format in ['excel', 'all']:
            excel_file = self.generate_excel_report(analysis_results)
            generated_files['excel'] = excel_file
            
        if format in ['html', 'all']:
            html_file = self.generate_html_report(analysis_results)
            generated_files['html'] = html_file
            
        if format in ['markdown', 'all']:
            md_file = self.generate_markdown_report(analysis_results)
            generated_files['markdown'] = md_file
        
        logger.info(f"Reports generated: {generated_files}")
        return generated_files
    
    def generate_json_report(self, analysis_results: Dict) -> str:
        """Generate JSON report"""
        
        filename = self.output_dir / f'analysis_report_{self.timestamp}.json'
        
        # Convert any dataframes to dictionaries
        json_data = self._prepare_for_json(analysis_results)
        
        with open(filename, 'w') as f:
            json.dump(json_data, f, indent=2, default=str)
        
        logger.info(f"JSON report saved to {filename}")
        return str(filename)
    
    def _prepare_for_json(self, data: Any) -> Any:
        """Prepare data for JSON serialization"""
        
        if isinstance(data, pd.DataFrame):
            return data.to_dict('records')
        elif isinstance(data, dict):
            return {k: self._prepare_for_json(v) for k, v in data.items()}
        elif isinstance(data, list):
            return [self._prepare_for_json(item) for item in data]
        elif isinstance(data, (np.integer, np.floating)):
            return float(data)
        elif isinstance(data, np.ndarray):
            return data.tolist()
        else:
            return data
    
    def generate_csv_reports(self, analysis_results: Dict) -> List[str]:
        """Generate CSV reports for tabular data"""
        
        csv_files = []
        
        # Duration statistics
        if 'duration_stats' in analysis_results:
            df = self._duration_stats_to_df(analysis_results['duration_stats'])
            filename = self.output_dir / f'duration_stats_{self.timestamp}.csv'
            df.to_csv(filename, index=False)
            csv_files.append(str(filename))
        
        # Outcome statistics
        if 'outcome_stats' in analysis_results:
            df = self._outcome_stats_to_df(analysis_results['outcome_stats'])
            filename = self.output_dir / f'outcome_stats_{self.timestamp}.csv'
            df.to_csv(filename, index=False)
            csv_files.append(str(filename))
        
        # Method comparison
        if 'method_comparison' in analysis_results:
            df = analysis_results['method_comparison']
            if isinstance(df, pd.DataFrame):
                filename = self.output_dir / f'method_comparison_{self.timestamp}.csv'
                df.to_csv(filename, index=False)
                csv_files.append(str(filename))
        
        # Post-breakout performance
        if 'post_breakout_performance' in analysis_results:
            df = self._performance_to_df(analysis_results['post_breakout_performance'])
            filename = self.output_dir / f'post_breakout_performance_{self.timestamp}.csv'
            df.to_csv(filename, index=False)
            csv_files.append(str(filename))
        
        logger.info(f"CSV reports saved: {csv_files}")
        return csv_files
    
    def _duration_stats_to_df(self, duration_stats: Dict) -> pd.DataFrame:
        """Convert duration statistics to DataFrame"""
        
        rows = []
        for method, stats in duration_stats.items():
            row = {
                'method': method,
                'count': stats.get('count', 0),
                'mean': stats.get('mean', 0),
                'median': stats.get('median', 0),
                'std': stats.get('std', 0),
                'min': stats.get('min', 0),
                'max': stats.get('max', 0),
                'q25': stats.get('q25', 0),
                'q75': stats.get('q75', 0)
            }
            rows.append(row)
        
        return pd.DataFrame(rows)
    
    def _outcome_stats_to_df(self, outcome_stats: Dict) -> pd.DataFrame:
        """Convert outcome statistics to DataFrame"""
        
        rows = []
        for method, stats in outcome_stats.items():
            row = {
                'method': method,
                'total_breakouts': stats.get('total_breakouts', 0),
                'up_breakouts': stats.get('up_breakouts', 0),
                'down_breakouts': stats.get('down_breakouts', 0),
                'up_ratio': stats.get('up_ratio', 0),
                'avg_max_gain': stats.get('avg_max_gain', 0),
                'success_rate': stats.get('success_rate', 0)
            }
            
            # Add outcome distribution
            distribution = stats.get('outcome_distribution', {})
            for outcome_class in ['K0', 'K1', 'K2', 'K3', 'K4', 'K5']:
                row[f'{outcome_class}_count'] = distribution.get(outcome_class, 0)
            
            rows.append(row)
        
        return pd.DataFrame(rows)
    
    def _performance_to_df(self, performance_data: Dict) -> pd.DataFrame:
        """Convert performance data to DataFrame"""
        
        rows = []
        for days, metrics in performance_data.items():
            row = {'days': days}
            row.update(metrics)
            rows.append(row)
        
        return pd.DataFrame(rows)
    
    def generate_excel_report(self, analysis_results: Dict) -> str:
        """Generate comprehensive Excel report with multiple sheets"""
        
        filename = self.output_dir / f'analysis_report_{self.timestamp}.xlsx'
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Add summary sheet
        self._add_summary_sheet(wb, analysis_results)
        
        # Add duration statistics sheet
        if 'duration_stats' in analysis_results:
            self._add_duration_sheet(wb, analysis_results['duration_stats'])
        
        # Add outcome statistics sheet
        if 'outcome_stats' in analysis_results:
            self._add_outcome_sheet(wb, analysis_results['outcome_stats'])
        
        # Add method comparison sheet
        if 'method_comparison' in analysis_results:
            self._add_comparison_sheet(wb, analysis_results['method_comparison'])
        
        # Add performance sheet
        if 'post_breakout_performance' in analysis_results:
            self._add_performance_sheet(wb, analysis_results['post_breakout_performance'])
        
        # Add optimal patterns sheet
        if 'optimal_patterns' in analysis_results:
            self._add_optimal_patterns_sheet(wb, analysis_results['optimal_patterns'])
        
        wb.save(filename)
        logger.info(f"Excel report saved to {filename}")
        return str(filename)
    
    def _add_summary_sheet(self, wb: Workbook, analysis_results: Dict) -> None:
        """Add summary sheet to Excel workbook"""
        
        ws = wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = "Consolidation Pattern Analysis Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')
        
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True)
        
        # Summary statistics
        row = 4
        summary = analysis_results.get('summary', {})
        
        ws[f'A{row}'] = "Key Metrics"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        metrics = [
            ('Total Patterns', summary.get('total_patterns', 0)),
            ('Unique Tickers', summary.get('unique_tickers', 0)),
            ('Overall Success Rate', f"{summary.get('overall_success_rate', 0):.1%}"),
            ('Exceptional Patterns (K4)', summary.get('exceptional_patterns', 0)),
            ('Failed Patterns (K5)', summary.get('failed_patterns', 0))
        ]
        
        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            row += 1
        
        # Key findings
        row += 1
        ws[f'A{row}'] = "Key Findings"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        findings = summary.get('key_findings', [])
        for finding in findings:
            ws[f'A{row}'] = finding
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _add_duration_sheet(self, wb: Workbook, duration_stats: Dict) -> None:
        """Add duration statistics sheet"""
        
        ws = wb.create_sheet("Duration Analysis")
        
        # Convert to DataFrame
        df = self._duration_stats_to_df(duration_stats)
        
        # Add headers
        headers = df.columns.tolist()
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)
        
        # Format columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = 15
    
    def _add_outcome_sheet(self, wb: Workbook, outcome_stats: Dict) -> None:
        """Add outcome statistics sheet"""
        
        ws = wb.create_sheet("Outcome Analysis")
        
        # Convert to DataFrame
        df = self._outcome_stats_to_df(outcome_stats)
        
        # Add headers
        headers = df.columns.tolist()
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)
        
        # Format percentage columns
        percentage_cols = ['up_ratio', 'success_rate']
        for col_name in percentage_cols:
            if col_name in headers:
                col_idx = headers.index(col_name) + 1
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value is not None:
                        cell.number_format = '0.0%'
    
    def _add_comparison_sheet(self, wb: Workbook, comparison_data: Any) -> None:
        """Add method comparison sheet"""
        
        ws = wb.create_sheet("Method Comparison")
        
        if isinstance(comparison_data, pd.DataFrame):
            df = comparison_data
        else:
            df = pd.DataFrame(comparison_data)
        
        # Add headers
        headers = df.columns.tolist()
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)
    
    def _add_performance_sheet(self, wb: Workbook, performance_data: Dict) -> None:
        """Add post-breakout performance sheet"""
        
        ws = wb.create_sheet("Post-Breakout Performance")
        
        # Convert to DataFrame
        df = self._performance_to_df(performance_data)
        
        # Add headers
        headers = df.columns.tolist()
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)
        
        # Format percentage columns
        for col_name in ['avg_gain', 'avg_loss', 'win_rate']:
            if col_name in headers:
                col_idx = headers.index(col_name) + 1
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value is not None:
                        cell.number_format = '0.0%' if col_name == 'win_rate' else '0.00'
    
    def _add_optimal_patterns_sheet(self, wb: Workbook, optimal_patterns: List[Dict]) -> None:
        """Add optimal patterns sheet"""
        
        ws = wb.create_sheet("Optimal Patterns")
        
        # Add headers
        ws['A1'] = "Optimal Pattern Characteristics"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:E1')
        
        row = 3
        ws['A3'] = "Metric"
        ws['B3'] = "Mean"
        ws['C3'] = "Median"
        ws['D3'] = "Optimal Range (Q25-Q75)"
        
        for col in range(1, 5):
            ws.cell(row=3, column=col).font = Font(bold=True)
            ws.cell(row=3, column=col).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        row = 4
        for pattern in optimal_patterns:
            ws[f'A{row}'] = pattern.get('metric', '')
            ws[f'B{row}'] = f"{pattern.get('mean', 0):.2f}"
            ws[f'C{row}'] = f"{pattern.get('median', 0):.2f}"
            
            opt_range = pattern.get('optimal_range', (0, 0))
            ws[f'D{row}'] = f"{opt_range[0]:.2f} - {opt_range[1]:.2f}"
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
    
    def generate_html_report(self, analysis_results: Dict) -> str:
        """Generate HTML report"""
        
        filename = self.output_dir / f'analysis_report_{self.timestamp}.html'
        
        html_template = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Consolidation Pattern Analysis Report</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; }
                h1 { color: #2c3e50; }
                h2 { color: #34495e; border-bottom: 2px solid #ecf0f1; padding-bottom: 5px; }
                table { border-collapse: collapse; width: 100%; margin: 20px 0; }
                th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                th { background-color: #3498db; color: white; }
                tr:nth-child(even) { background-color: #f2f2f2; }
                .metric { font-weight: bold; }
                .success { color: green; }
                .failure { color: red; }
                .summary-box { background-color: #ecf0f1; padding: 15px; border-radius: 5px; margin: 20px 0; }
            </style>
        </head>
        <body>
            <h1>Consolidation Pattern Analysis Report</h1>
            <p><em>Generated: {{ timestamp }}</em></p>
            
            <div class="summary-box">
                <h2>Executive Summary</h2>
                <p><span class="metric">Total Patterns:</span> {{ summary.total_patterns }}</p>
                <p><span class="metric">Success Rate:</span> {{ "%.1f" | format(summary.overall_success_rate * 100) }}%</p>
                <p><span class="metric">Exceptional Patterns:</span> {{ summary.exceptional_patterns }}</p>
                <p><span class="metric">Failed Patterns:</span> {{ summary.failed_patterns }}</p>
            </div>
            
            {% if duration_stats %}
            <h2>Duration Analysis</h2>
            <table>
                <tr>
                    <th>Method</th>
                    <th>Count</th>
                    <th>Mean</th>
                    <th>Median</th>
                    <th>Min</th>
                    <th>Max</th>
                </tr>
                {% for method, stats in duration_stats.items() %}
                <tr>
                    <td>{{ method }}</td>
                    <td>{{ stats.count }}</td>
                    <td>{{ "%.1f" | format(stats.mean) }}</td>
                    <td>{{ "%.1f" | format(stats.median) }}</td>
                    <td>{{ stats.min }}</td>
                    <td>{{ stats.max }}</td>
                </tr>
                {% endfor %}
            </table>
            {% endif %}
            
            {% if key_findings %}
            <h2>Key Findings</h2>
            <ul>
                {% for finding in key_findings %}
                <li>{{ finding }}</li>
                {% endfor %}
            </ul>
            {% endif %}
        </body>
        </html>
        """
        
        template = Template(html_template)
        
        html_content = template.render(
            timestamp=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            summary=analysis_results.get('summary', {}),
            duration_stats=analysis_results.get('duration_stats', {}),
            key_findings=analysis_results.get('summary', {}).get('key_findings', [])
        )
        
        with open(filename, 'w') as f:
            f.write(html_content)
        
        logger.info(f"HTML report saved to {filename}")
        return str(filename)
    
    def generate_markdown_report(self, analysis_results: Dict) -> str:
        """Generate Markdown report"""
        
        filename = self.output_dir / f'analysis_report_{self.timestamp}.md'
        
        md_content = f"""# Consolidation Pattern Analysis Report

*Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}*

## Executive Summary

"""
        
        summary = analysis_results.get('summary', {})
        md_content += f"""
- **Total Patterns:** {summary.get('total_patterns', 0)}
- **Unique Tickers:** {summary.get('unique_tickers', 0)}
- **Overall Success Rate:** {summary.get('overall_success_rate', 0):.1%}
- **Exceptional Patterns (K4):** {summary.get('exceptional_patterns', 0)}
- **Failed Patterns (K5):** {summary.get('failed_patterns', 0)}

"""
        
        # Duration Statistics
        if 'duration_stats' in analysis_results:
            md_content += "## Duration Analysis\n\n"
            md_content += "| Method | Count | Mean | Median | Min | Max |\n"
            md_content += "|--------|-------|------|--------|-----|-----|\n"
            
            for method, stats in analysis_results['duration_stats'].items():
                md_content += f"| {method} | {stats.get('count', 0)} | "
                md_content += f"{stats.get('mean', 0):.1f} | {stats.get('median', 0):.1f} | "
                md_content += f"{stats.get('min', 0)} | {stats.get('max', 0)} |\n"
            
            md_content += "\n"
        
        # Outcome Statistics
        if 'outcome_stats' in analysis_results:
            md_content += "## Outcome Analysis\n\n"
            md_content += "| Method | Total Breakouts | Up | Down | Success Rate |\n"
            md_content += "|--------|-----------------|-----|------|-------------|\n"
            
            for method, stats in analysis_results['outcome_stats'].items():
                md_content += f"| {method} | {stats.get('total_breakouts', 0)} | "
                md_content += f"{stats.get('up_breakouts', 0)} | {stats.get('down_breakouts', 0)} | "
                md_content += f"{stats.get('success_rate', 0):.1%} |\n"
            
            md_content += "\n"
        
        # Key Findings
        if 'key_findings' in summary:
            md_content += "## Key Findings\n\n"
            for finding in summary['key_findings']:
                md_content += f"- {finding}\n"
            md_content += "\n"
        
        # Method Comparison
        if 'method_comparison' in analysis_results:
            md_content += "## Method Comparison\n\n"
            comparison = analysis_results['method_comparison']
            
            if isinstance(comparison, pd.DataFrame) and not comparison.empty:
                md_content += comparison.to_markdown(index=False)
                md_content += "\n\n"
        
        # Optimal Patterns
        if 'optimal_patterns' in analysis_results:
            md_content += "## Optimal Pattern Characteristics\n\n"
            md_content += "| Metric | Mean | Median | Optimal Range |\n"
            md_content += "|--------|------|--------|---------------|\n"
            
            for pattern in analysis_results['optimal_patterns']:
                opt_range = pattern.get('optimal_range', (0, 0))
                md_content += f"| {pattern.get('metric', '')} | "
                md_content += f"{pattern.get('mean', 0):.2f} | {pattern.get('median', 0):.2f} | "
                md_content += f"{opt_range[0]:.2f} - {opt_range[1]:.2f} |\n"
            
            md_content += "\n"
        
        with open(filename, 'w') as f:
            f.write(md_content)
        
        logger.info(f"Markdown report saved to {filename}")
        return str(filename)
    
    def create_summary_dashboard(self, analysis_results: Dict) -> pd.DataFrame:
        """Create summary dashboard as DataFrame"""
        
        dashboard_data = []
        
        # Overall metrics
        summary = analysis_results.get('summary', {})
        dashboard_data.append({
            'Category': 'Overall',
            'Metric': 'Total Patterns',
            'Value': summary.get('total_patterns', 0)
        })
        dashboard_data.append({
            'Category': 'Overall',
            'Metric': 'Success Rate',
            'Value': f"{summary.get('overall_success_rate', 0):.1%}"
        })
        
        # Method comparison
        if 'method_comparison' in analysis_results:
            comparison = analysis_results['method_comparison']
            if isinstance(comparison, pd.DataFrame):
                best_method = comparison.loc[comparison['success_rate'].idxmax()]
                dashboard_data.append({
                    'Category': 'Best Method',
                    'Metric': 'Method Name',
                    'Value': best_method.get('method', 'N/A')
                })
                dashboard_data.append({
                    'Category': 'Best Method',
                    'Metric': 'Success Rate',
                    'Value': f"{best_method.get('success_rate', 0):.1%}"
                })
        
        # Risk metrics
        if 'risk_metrics' in analysis_results:
            risk = analysis_results['risk_metrics']
            dashboard_data.append({
                'Category': 'Risk',
                'Metric': 'Failure Rate',
                'Value': f"{risk.get('failure_rate', 0):.1%}"
            })
            
            if 'risk_reward_ratios' in risk:
                dashboard_data.append({
                    'Category': 'Risk',
                    'Metric': 'Risk/Reward Ratio',
                    'Value': f"{risk['risk_reward_ratios'].get('ratio', 0):.2f}"
                })
        
        return pd.DataFrame(dashboard_data)